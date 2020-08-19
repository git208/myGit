import openpyxl
import os
import json
import re

path_list = []

#不安全的递归查询
def search_unsafe(path=".", name=""):
    for item in os.listdir(path):
        item_path = str(path+'/'+item)
        if os.path.isdir(item_path):
            search_unsafe(item_path, name)
        elif os.path.isfile(item_path):
            if name in item:
                global path_list
                print(item_path)
                path_list.append(item_path)

#模糊查询指定路径下所有匹配文件的绝对路径，返回查到的路径列表
def search_file(path=".", name=""):
    global path_list
    path_list = []
    print('.......开始在 %s 下模糊查询 %s'%(path,name))
    search_unsafe(path, name)
    return path_list

#排序
def mySort(list):
    sort_list = []
    for temp in list:
        sort_list.append(temp['s'])
    sort_list.sort()
    return sort_list

#根据板块名查其下所有分类文件路径,p3：格式附加名
def search_bl(path=".", bankuaiName="",name_additional = ''):
    print('------------尝试查询板块【%s】目录文件的所在路径------------'%(bankuaiName))
    pa = search_file(path,bankuaiName)
    print('板块%s目录文件的所在路径为%s' % (bankuaiName,pa))
    filepath_list = []
    with open(file=pa[0], mode='r', encoding='utf-8') as a:
        temp_list = json.loads(a.read())
        print('尝试读取%s\n读取结果:%s'%(pa[0],temp_list))
        #对静态板块目录文件中板块下分类进行排序
        sort_list = mySort(temp_list)
        print('取出分类代码并进行排序后：\n%s' % (sort_list))
        for tp in sort_list:
            aa = search_file(path,name_additional + tp)
            print(aa)
            filepath_list = filepath_list + aa
    print('有序获得获得%s下所有分类文件路径：\n%s'%(bankuaiName,filepath_list))
    return temp_list,filepath_list

#处理函数,p1:取得的文件所在的本地路径，p2:板块名，p3：格式附加名，p4是否生成没啥用的txt文件
def chuli(path,bankuaiName,name_additional = '',create_file = True):
    stock_list = []
    print('查询是否存在【%s】板块目录静态文件'%(bankuaiName))
    if search_file(path,bankuaiName):
        print('板块目录文件【%s】文件存在' % (bankuaiName))
        print('------------开始查询板块【%s】下所有分类文件所在的路径------------'%(bankuaiName))
        paths = search_bl(path,bankuaiName,name_additional)
        if create_file:
            _ = open(file='zhenghe.txt', mode='w', encoding='utf-8')
            _.close()
        for pt in paths[1]:
            with open(file=pt, mode='r', encoding='utf-8') as b:
                hade_dict = {}
                for tpl in paths[0]:
                    if re.sub('\D','',os.path.basename(pt)) == re.sub('\D','',tpl['s']):
                        hade_dict = tpl
                temp_file = hade_dict['s']+','+ hade_dict['ns']+','+ hade_dict['n']+','+\
                            b.read().replace('{"s":"','').replace('"}','').replace('[','').replace(']','')
                temp_file_list = temp_file.split(',')
                #对各版块分类下股票代码进行排序
                temp_file_list_front = temp_file_list[:3]
                temp_file_list_back = temp_file_list[3:]
                temp_file_list_back.sort()
                temp_file_list = temp_file_list_front + temp_file_list_back
                stock_list = stock_list + temp_file_list
                if create_file:
                    with open(file='zhenghe.txt', mode='a', encoding='utf-8') as integration_file:
                        integration_file.write(temp_file)
        print('静态文件经过处理后的数据：\n%s' % (stock_list))
    else:
        print('板块目录文件%s文件不存在' % (bankuaiName))
    return stock_list

#向excle中插入整理后的数据
def excle(Integrated_list,name,xlsx):
    xlsx.create_sheet(title=name)
    print('创建了一个新的sheet页：%s' % (name))
    sheet = xlsx[name]
    print('向sheet页【%s】中插入数据' % (name))
    sheet['b1'] = name
    for i in range(len(Integrated_list)):
        sheet['b' + str(2 + i)] = Integrated_list[i]
        sheet['c' + str(1 + i)] = '=EXACT(A' + str(1 + i) + ', B' + str(1 + i) + ')'
    if len(Integrated_list) > 0:
        sheet['c' + str(len(Integrated_list)+1)] = '=EXACT(A' + str(len(Integrated_list)+1) + ', B' + str(len(Integrated_list)+1) + ')'
    print('sheet页【%s】数据插入完毕' % (name))

#循环插入excle，p1：静态文件所在位置; p2：想要生成的excle名; p3：板块名json
def excle_more(path,resultName,bankuaiNames,create_file = True):
    print('建立一个新的工作薄')
    if os.path.exists(resultName):
        os.remove(resultName)
        excl = openpyxl.Workbook()
    else:
        excl = openpyxl.Workbook()
    for key_name,value_additionalName in bankuaiNames.items():
        print('****************开始处理板块%s相关文件****************'%(key_name))
        Integrated_list = chuli(path, key_name, value_additionalName, create_file)
        print('****************板块%s相关文件处理完毕，开始插入excle****************' % (key_name))
        excle(Integrated_list,key_name,excl)
    print('所有数据插入完毕，保存工作薄为【%s】'%(resultName))
    excl.save(resultName)

#向excle中插入从行情监控系统中粘贴过来的数据
def excle_monitor(path,bankuaiName,resultName):
    print('打开工作簿：%s' % (resultName))
    xlsx = openpyxl.load_workbook(resultName)
    for key,_ in bankuaiName.items():
        print('查询是否存在从行情监控系统得到的 %s 板块文件' % (key))
        if search_file(path,key):
            path_name = search_file(path, key)
            with open(file=path_name[0], mode='r', encoding='utf-8') as a:
                temp_file = a.read().split('')
                hade = temp_file[:1]
                stock_list = []
                temp_list = temp_file[1].split('')
                for tp in temp_list:
                    temp_file_list = tp.split(',')
                    temp_file_list_front = temp_file_list[:3]
                    temp_file_list_back = temp_file_list[3:]
                    temp_file_list_back.sort()
                    temp_file_list = temp_file_list_front + temp_file_list_back
                    stock_list = stock_list + temp_file_list
                stock_list = hade + stock_list
                print('经过排序处理后的数据：\n%s'%(stock_list))
                sheet = xlsx[key]
                print('向sheet页【%s】中插入整理后的数据' % (key))
                for i in range(len(stock_list)):
                    sheet['a' + str(1 + i)] = stock_list[i]
                print('sheet页【%s】数据插入完毕' % (key))
    print('所有行情监控系统数据插入完毕，保存工作薄【%s】' % (resultName))
    xlsx.save(resultName)